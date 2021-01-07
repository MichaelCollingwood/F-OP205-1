""" Training Input Data """

import numpy as np
import matplotlib.pyplot as plt
import os
import torch
import random
import json
import librosa
from tqdm import tqdm
from scipy.signal import find_peaks

from PeakBuildModels import BuildModels_p
from FeatureExtractor import FeatureExtractor
from InputFormatting import ImageToTorchFormat
from DataStorage import OpenCache_llambda
from DataManipulation import ButterLowpassFilter
from PeaksMatrix import GetPeakVector
from openpyxl import load_workbook
from Display import PlotSpecAndFunc


def ManuallyLabelPeaks(model, syll, window, step, nperseg):
    """
    Label Peaks: Syll/Non-Syll
    """
    nMS = GetNormativeManualScoring()
    keys = [key for key in nMS.keys() if (nMS[key][0][:2] == syll) and (nMS[key][0] != "pa ta ka")]
    random.shuffle(keys)
    
    peakMatrix = []; output_p = []
    
    for key in keys[:1]:
        manuallyScoredData = nMS[key]

        """ Get -> FFT -> Window -> Format """
        signal, rate = librosa.load("../audio_data/pataka-norms-audio/"+key+".wav")
        
        llambda_samples = OpenCache_llambda(nperseg)
        np.random.shuffle(llambda_samples)
        
        Times = np.array([]); Ys = np.array([])
        for llambda in tqdm(llambda_samples[:400]):
            extractor = FeatureExtractor(signal, rate)
            extractor.GetSpectrogram(nperseg)
            extractor.SetFeatures("spectrogram")

            extractor.Window(int(llambda*window), step)
            extractor.SqueezeWindows(window)
            times, X = extractor.features
            X = ImageToTorchFormat(X)
    
            """ CNN Envelope """        
            with torch.no_grad():
                output = model(X)
            output = np.array(output[:,1])
            
            Times = np.concatenate((Times, times))
            Ys = np.concatenate((Ys, np.exp(output)))
        
        Ys = Ys[np.argsort(Times)]
        times = Times[np.argsort(Times)]
        y = ButterLowpassFilter(Ys, 0.006)
        y -= np.min(y)
        y /= np.max(y)
        
        """ Peaks """
        peaks, peakProperties = find_peaks(y, height=0)
        peaksVectors = GetPeakVector(peaks, y, times)
        
        for i, peak in enumerate(peaks):
            peakMatrix.append(peaksVectors[i])
                
            PlotSpecAndFunc(extractor.spectrogram[0], extractor.spectrogram[1], extractor.spectrogram[2], times, y, np.array([peak]), manuallyScoredData)
            
            if (input('is syllable?[y/]') == 'y'):
                output_p.append(1)
            else:
                output_p.append(0)

    if (os.path.exists("output.json")):
        with open("output.json", "r") as read_file:
            Y = json.load(read_file)
        for element in output_p:
            Y.append(element)
        with open("output.json", "w") as fp:
            json.dump(Y, fp)
    else:
        with open("output.json", "w") as fp:
            json.dump(output_p, fp)

    if (os.path.exists("peakMatrix.json")):
        with open("peakMatrix.json", "r") as read_file:
            X = json.load(read_file)
        for row in peakMatrix:
            X.append(row)
        with open("peakMatrix.json", "w") as fp:
            json.dump(X, fp)
    else:
        with open("peakMatrix.json", "w") as fp:
            json.dump(peakMatrix, fp)
        
        # train 1D CNN and save it
        BuildModels_p()        


def GetNormativeManualScoring():
    normativeManualScoring = load_workbook(filename="../audio_data/NormativeManualScoring.xlsx")
    sheet = normativeManualScoring.active
    
    i = 2
    normativeManualScoring = {}
    while (sheet["B"+str(i)].value != None):
        utteranceId = sheet["B"+str(i)].value
        
        targetSyllable = sheet["A"+str(i)].value
        onset = sheet["D"+str(i)].value
        offset = sheet["E"+str(i)].value
        repetitions = sheet["F"+str(i)].value
        
        if (repetitions != None):
            normativeManualScoring[utteranceId] = [targetSyllable, onset, offset, repetitions]
        
        i += 1
    
    return normativeManualScoring

def GetLabels(key):
    normativeManualScoring = load_workbook(filename="../audio_data/NormativeManualScoring.xlsx")
    sheet = normativeManualScoring.active
    
    i = 2
    normativeManualScoring = {}
    while ((sheet["B"+str(i)].value != key) and (sheet["B"+str(i)].value != None)):
        i += 1
    
    targetSyllable = sheet["A"+str(i)].value
    onset = sheet["D"+str(i)].value
    offset = sheet["E"+str(i)].value
    repetitions = sheet["F"+str(i)].value
    
    return [targetSyllable, onset, offset, repetitions]


def TrimUtterance(signal, onset, offset, fs = 16000):
    onsetIndex = int(onset * fs)
    offsetIndex = int(offset * fs)
    
    return signal[onsetIndex:offsetIndex]



def GetInflexionTimes(y, locations, times, plot=False):
    # find inflexion points encolsing each peak
    
    yd2 = np.gradient(np.gradient(y))
    infls = np.where(np.diff(np.sign(yd2)))[0]
    
    if (plot):
        # demonstrate
        plt.plot(y)
        plt.vlines(infls, max(y), min(y), color='black')
        plt.xlim(0,2000)
        plt.show()
    
    infls = np.concatenate((np.array([0]), infls, np.array([len(y)])))
    
    inflextionPts = []
    for location in locations:
        z = infls - location
        z1 = np.array([zi if zi > 0 else np.max(np.abs(z)) for zi in z])
        z2 = np.array([-zi if zi < 0 else np.max(np.abs(z)) for zi in z])
        upperInflPt = np.argmin(z1)
        lowerInflPt = np.argmin(z2)
        inflextionPts.append((infls[lowerInflPt],infls[upperInflPt]))
    
    times_ = np.concatenate((np.array([times[0]]), times, np.array([times[-1]])))
    return [list(row) for row in times_[np.array(inflextionPts)]]



def ThinInflexionTimes(infl_times, factor):
    
    infl_times_new = []
    for interval in infl_times:
        duration = interval[1] - interval[0]
        interval = [interval[0] + duration*factor/2, interval[1] - duration*factor/2]
        infl_times_new.append(interval)
    
    return infl_times_new



def FindGroundTruth(times_new, inflexion_points):
    GT = []
    for ti in times_new:
        syll = 0
        for inflection_couple in inflexion_points:
            if ((ti > inflection_couple[0]) and (ti < inflection_couple[1])):
                syll = 1
        GT.append(syll)
    return np.array(GT)